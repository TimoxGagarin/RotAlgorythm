from openpyxl import Workbook
from cube import Cube
from copy import copy


def num_to_chars(num):
    str = ''
    p = 0
    while num >= 26:
        p = 0
        temp = num
        while temp >= 26:
            temp //= 26
            p += 1
        str += chr(64 + temp)
        num -= temp*pow(26, p)
    if num != 0:
        str += chr(64 + num)
    return str


def has_k_x(cub, k):
    cnt = 0
    for el in cub.string:
        if el == 'x':
            cnt += 1
    return cnt >= k


class RotAlgos:

    def __init__(self, L, N):
        self.L = L
        self.N = N

        # original coating
        self.C0 = L | N
        # simple implicants
        self.Z = set()

        self.workbook = Workbook()

    def run(self):
        sheet = self.workbook['Sheet']

        self.Z = self.find_simple_implicants()
        print(f"Простые импликанты: {self.Z}")
        E = self.find_L_extr()
        print(f"L-экстремали: {E}")
        Z_res = E

        ost = self.find_non_covered_cubes(E)
        print(f"Кубы, не покрываемые L-экстремалями: {ost}")
        L_ost = self.Z - E

        # coming soon...
        solution = self.find_other_cubes(L_ost, ost)

    def find_simple_implicants(self):
        sheet = self.workbook['Sheet']
        sheet.title = "C0 m C0"

        Z_res = set()
        C = self.C0
        k = 0

        while True:
            Z = copy(C)
            A = set()
            B = set()

            # generate table grafs
            sheet["A1"] = sheet.title
            C_list = list(C)
            for i in range(0, len(C_list)):
                sheet[f"{num_to_chars(i + 2)}1"] = C_list[i].string
                sheet[f"A{i + 2}"] = C_list[i].string
            sheet[f"A{len(C_list)+2}"] = f'A{k+1}'
            k += 1

            A_list = list()
            for i in range(len(C_list)):
                A_list.append(set())
            # multi cubes in table
            for i in range(len(C_list)):
                for j in range(i + 1):
                    if i == j:
                        sheet[f"{num_to_chars(j + 2)}{i + 2}"] = '-'
                        continue
                    val = C_list[i].multi(C_list[j])
                    if val is None or not has_k_x(val, k):
                        sheet[f"{num_to_chars(j + 2)}{i + 2}"] = "-"
                    else:
                        sheet[f"{num_to_chars(j + 2)}{i + 2}"] = val.string
                        A.add(val)
                        A_list[j].add(val)
                        Z -= {C_list[i], C_list[j]}
            for i in range(len(A_list)):
                cell_text = ''
                for el in A_list[i]:
                    cell_text += f"{el}\r\n"
                sheet[f"{num_to_chars(i+2)}{len(C_list) + 2}"] = cell_text
            B = C - Z
            B_copy = copy(B)

            Z_res |= Z
            for el in B:
                if not has_k_x(el, k):
                    B_copy -= {el}
            B = B_copy
            C = A | B
            if C == set():
                break
            sheet = self.workbook.create_sheet(f"C{k} m C{k}")
        self.workbook.save(filename="result.xlsx")

        return Z_res

    def find_L_extr(self):
        # generate table grafs
        sheet = self.workbook.create_sheet("z#(Z-z)")
        L_extr = set()
        Z_list = list(self.Z)

        sheet["A1"] = sheet.title
        for i in range(0, len(self.Z)):
            sheet[f"{num_to_chars(i + 2)}1"] = Z_list[i].string
            sheet[f"A{i + 2}"] = Z_list[i].string

        # find L-extremals
        to_use_list = list()
        for i in range(0, len(Z_list)):
            to_use_list.append({Z_list[i]})

        # difference and fill the table
        for i in range(len(self.Z)):
            for j in range(len(self.Z)):
                res = set()
                if i == j:
                    sheet[f"{num_to_chars(j + 2)}{i + 2}"] = '-'
                    continue

                if to_use_list[j] != set():
                    for el in to_use_list[j]:
                        res |= el.difference(Z_list[i])

                cell_text = ''
                to_use_list[j] = res

                for el in to_use_list[j]:
                    cell_text += f"{el}\r\n"
                sheet[f"{num_to_chars(j + 2)}{i + 2}"] = cell_text
        self.workbook.save(filename="result.xlsx")

        # find results of difference and potencial L-extremales
        after_diffrence = set()
        for i in range(len(to_use_list)):
            after_diffrence |= to_use_list[i]
            if to_use_list[i] != set():
                L_extr.add(Z_list[i])

        if L_extr != set() and self.N == set():
            return L_extr
        elif L_extr == set():
            for e in Z_list:
                for i in range(1, len(e)):
                    for el in Z_list:
                        if len(el.find_all('x')) == i:
                            self.Z.remove(el)
                            return self.find_L_extr()
                break

        # generate table grafs
        sheet = self.workbook.create_sheet("L & (Z-z)")
        E = copy(L_extr)

        after_diffrence_list = list(after_diffrence)
        L_list = list(self.L)

        sheet["A1"] = sheet.title
        for i in range(0, len(L_list)):
            sheet[f"{num_to_chars(i + 2)}1"] = L_list[i].string
        for i in range(0, len(after_diffrence_list)):
            sheet[f"A{i + 2}"] = after_diffrence_list[i].string

        # intersection and fill the table
        for i in range(len(after_diffrence_list)):
            is_cover_N = True
            for j in range(len(L_list)):
                res = L_list[j].intersection(after_diffrence_list[i])
                if res is not None:
                    is_cover_N = False
                    sheet[f"{num_to_chars(j + 2)}{i + 2}"] = str(res)
                else:
                    sheet[f"{num_to_chars(j + 2)}{i + 2}"] = '-'
            if is_cover_N:
                for el_set in to_use_list:
                    el_set -= {after_diffrence_list[i]}

        # remove L-extremals that cover only N
        for i in range(len(to_use_list)):
            if to_use_list[i] == set():
                L_extr -= {Z_list[i]}

        workbook.save(filename="result.xlsx")
        return L_extr

    def find_non_covered_cubes(self, E):
        # generate table grafs
        sheet = self.workbook.create_sheet("L # E")

        E_list = list(E)
        L_list = list(self.L)

        sheet["A1"] = sheet.title
        for i in range(0, len(self.L)):
            sheet[f"{num_to_chars(i + 2)}1"] = L_list[i].string
        for i in range(0, len(E)):
            sheet[f"A{i + 2}"] = E_list[i].string

        to_use_list = list()
        for i in range(0, len(L_list)):
            to_use_list.append({L_list[i]})

        # difference and fill the table
        for i in range(len(E_list)):
            for j in range(len(L_list)):
                res = set()

                if to_use_list[j] != set():
                    for el in to_use_list[j]:
                        res |= el.difference(E_list[i])

                cell_text = ''
                to_use_list[j] = res

                for el in to_use_list[j]:
                    cell_text += f"{el}\r\n"
                sheet[f"{num_to_chars(j + 2)}{i + 2}"] = cell_text if to_use_list[j] != set() else '-'
        self.workbook.save(filename="result.xlsx")

        # find non-covered cubes
        after_diffrence = set()
        for i in range(len(to_use_list)):
            after_diffrence |= to_use_list[i]

        return after_diffrence

    def find_other_cubes(self, L_ost, ost):
        # generate table grafs
        sheet = self.workbook.create_sheet("L' & Z^")

        ost_list = list(ost)
        L_ost_list = list(L_ost)

        sheet["A1"] = sheet.title
        for i in range(0, len(ost)):
            sheet[f"{num_to_chars(i + 2)}1"] = ost_list[i].string
        for i in range(0, len(L_ost)):
            sheet[f"A{i + 2}"] = L_ost_list[i].string

        covered_cubes_list = list()
        # intersection and fill the table
        for i in range(len(L_ost)):
            covered_cubes_list.append(list())
            for j in range(len(ost)):
                res = ost_list[j].intersection(L_ost_list[i])
                if res is not None:
                    covered_cubes_list[i].append(res)
                sheet[f"{num_to_chars(j + 2)}{i + 2}"] = str(res) if res is not None else '-'
        self.workbook.save(filename="result.xlsx")
        # to cover
        print(ost_list)
        # cover with
        print(L_ost_list)
        # covered by cubes
        print(covered_cubes_list)

        # coming soon...
        return None